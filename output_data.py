import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from collections import defaultdict
from datetime import datetime

class TimetableExporter:
    def __init__(self):
        self.saved_data_path = os.path.join(os.path.dirname(__file__), 'data', 'timetable_data.json')
        self.output_dir = os.path.join(os.path.dirname(__file__), 'output_data')
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Keywords to identify SST (engineering) groups
        self.sst_keywords = [
            'engineering', 'eng', 'computer science', 'software engineering', 'data science',
            'mechatronics', 'electrical', 'mechanical', 'csc', 'sen', 'data', 'ds'
        ]
        
        # Time slots
        self.time_slots = [
            "9:00-9:50", "10:00-10:50", "11:00-11:50", "12:00-12:50", 
            "1:00-1:50", "2:00-2:50", "3:00-3:50", "4:00-4:50", "5:00-5:50"
        ]
        
        # Days of the week
        self.days = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

    def load_saved_timetable_data(self):
        """Load the latest saved timetable data"""
        try:
            if os.path.exists(self.saved_data_path):
                with open(self.saved_data_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                print(f"✅ Loaded saved timetable data: {len(data)} groups")
                return data
            else:
                print("❌ No saved timetable data found")
                return None
        except Exception as e:
            print(f"❌ Error loading saved data: {e}")
            return None

    def is_sst_group(self, group_name):
        """Check if a student group belongs to SST (engineering) based on keywords"""
        group_name_lower = group_name.lower()
        return any(keyword in group_name_lower for keyword in self.sst_keywords)

    def extract_main_program_name(self, group_name):
        """Extract the main program name from student group name"""
        # Remove year information and stream information
        # Examples: "Computer Science - Year 1" -> "Computer Science"
        #          "Electrical Engineering - Year 2/Stream 1" -> "Electrical Engineering"
        
        # First check for hyphen
        if ' - ' in group_name:
            main_name = group_name.split(' - ')[0]
        else:
            # If no hyphen, look for the word "year" (case insensitive)
            match = re.search(r'\s+year\s+\d+', group_name, re.IGNORECASE)
            if match:
                main_name = group_name[:match.start()]
            else:
                main_name = group_name
        
        # Remove any stream information that might be in the main name
        main_name = re.sub(r'/Stream \d+', '', main_name)
        main_name = re.sub(r'Stream \d+', '', main_name)
        
        return main_name.strip()

    def extract_lecturer_info(self, cell_content):
        """Extract lecturer information from cell content"""
        if not cell_content or cell_content in ["FREE", "BREAK", ""]:
            return None
        
        # Cell format: Course Code\nRoom Name\nFaculty
        lines = cell_content.split('\n')
        if len(lines) >= 3:
            return {
                'course_code': lines[0].strip(),
                'room': lines[1].strip(),
                'faculty': lines[2].strip()
            }
        return None

    def create_combined_program_sheet(self, wb, sheet_name, groups_data):
        """Create a combined timetable sheet for multiple student groups of the same program"""
        ws = wb.create_sheet(title=sheet_name)
        
        current_row = 1
        
        for group_idx, group_data in enumerate(groups_data):
            full_group_name = group_data['student_group']['name']
            
            # Add group name header
            group_header_cell = ws.cell(row=current_row, column=1, value=f"{full_group_name} timetable")
            group_header_cell.font = Font(bold=True, size=14, color="FFFFFF")
            group_header_cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            group_header_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Merge cells for group header (across all days + time column)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(self.days) + 1)
            current_row += 1
            
            # Create header row with TIME and days
            headers = ["TIME"] + self.days
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += 1
            
            # Fill in the timetable data
            timetable_rows = group_data['timetable']
            
            for row_idx, row_data in enumerate(timetable_rows):
                # Add time slot
                time_cell = ws.cell(row=current_row, column=1, value=self.time_slots[row_idx])
                time_cell.font = Font(bold=True, color="FFFFFF")
                time_cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
                time_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Add class data for each day
                for day_idx in range(len(self.days)):
                    col = day_idx + 2  # Start from column 2 (after TIME column)
                    
                    if day_idx + 1 < len(row_data):  # Skip time column in source data
                        cell_content = row_data[day_idx + 1]
                        
                        if cell_content and cell_content not in ["FREE", ""]:
                            # Parse cell content
                            lines = cell_content.split('\n')
                            if len(lines) >= 3:
                                course_code = lines[0]
                                room = lines[1]
                                faculty = lines[2]
                                
                                # Format as: Course Code, Room, Faculty
                                display_text = f"{course_code}\n{room}\n{faculty}"
                            else:
                                display_text = cell_content
                        else:
                            display_text = ""
                        
                        cell = ws.cell(row=current_row, column=col, value=display_text)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        # Apply border
                        border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = border
                
                current_row += 1
            
            # Add spacing between student groups (except for the last one)
            if group_idx < len(groups_data) - 1:
                current_row += 2
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column].width = adjusted_width
        
        # Set row heights for better visibility
        for row in range(1, current_row):
            ws.row_dimensions[row].height = 50

    def create_student_group_timetable_sheet(self, wb, sheet_name, group_data):
        """Create a timetable sheet for a student group"""
        ws = wb.create_sheet(title=sheet_name)
        
        # Create header row with TIME and days
        headers = ["TIME"] + self.days
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Fill in the timetable data
        timetable_rows = group_data['timetable']
        
        for row_idx, row_data in enumerate(timetable_rows):
            excel_row = row_idx + 2  # Start from row 2 (after header)
            
            # Add time slot
            time_cell = ws.cell(row=excel_row, column=1, value=self.time_slots[row_idx])
            time_cell.font = Font(bold=True, color="FFFFFF")
            time_cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            time_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add class data for each day
            for day_idx in range(len(self.days)):
                col = day_idx + 2  # Start from column 2 (after TIME column)
                
                if day_idx + 1 < len(row_data):  # Skip time column in source data
                    cell_content = row_data[day_idx + 1]
                    
                    if cell_content and cell_content not in ["FREE", ""]:
                        # Parse cell content
                        lines = cell_content.split('\n')
                        if len(lines) >= 3:
                            course_code = lines[0]
                            room = lines[1]
                            faculty = lines[2]
                            
                            # Format as: Course Code, Room, Faculty
                            display_text = f"{course_code}\n{room}\n{faculty}"
                        else:
                            display_text = cell_content
                    else:
                        display_text = ""
                    
                    cell = ws.cell(row=excel_row, column=col, value=display_text)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Apply border
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column].width = adjusted_width
        
        # Set row heights for better visibility
        for row in range(1, len(timetable_rows) + 2):
            ws.row_dimensions[row].height = 60

    def create_lecturer_timetable_sheet(self, wb, lecturer_name, lecturer_schedule):
        """Create a timetable sheet for a lecturer"""
        # Clean lecturer name for sheet title
        safe_name = re.sub(r'[^\w\s-]', '', lecturer_name)[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=safe_name)
        
        # Create header row
        headers = ["TIME"] + self.days
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Initialize empty schedule grid
        schedule_grid = {}
        for time_idx in range(len(self.time_slots)):
            schedule_grid[time_idx] = {}
            for day_idx in range(len(self.days)):
                schedule_grid[time_idx][day_idx] = ""
        
        # Fill in the lecturer's classes
        for class_info in lecturer_schedule:
            time_idx = class_info['time_slot']
            day_idx = class_info['day']
            
            display_text = f"{class_info['course_code']}\n{class_info['room']}\n{class_info['student_group']}"
            schedule_grid[time_idx][day_idx] = display_text
        
        # Fill in the Excel sheet
        for time_idx in range(len(self.time_slots)):
            excel_row = time_idx + 2
            
            # Add time slot
            time_cell = ws.cell(row=excel_row, column=1, value=self.time_slots[time_idx])
            time_cell.font = Font(bold=True, color="FFFFFF")
            time_cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            time_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add classes for each day
            for day_idx in range(len(self.days)):
                col = day_idx + 2
                cell_content = schedule_grid[time_idx][day_idx]
                
                cell = ws.cell(row=excel_row, column=col, value=cell_content)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Apply border
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = border
        
        # Adjust column widths and row heights
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column].width = adjusted_width
        
        for row in range(1, len(self.time_slots) + 2):
            ws.row_dimensions[row].height = 60

    def export_sst_timetables(self):
        """Export SST (engineering) timetables"""
        data = self.load_saved_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Group SST student groups by main program
            sst_programs = defaultdict(list)
            
            for group_data in data:
                group_name = group_data['student_group']['name']
                if self.is_sst_group(group_name):
                    main_program = self.extract_main_program_name(group_name)
                    sst_programs[main_program].append(group_data)
            
            if not sst_programs:
                return False, "No SST student groups found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each unique program (one sheet per main program)
            for program_name, groups in sst_programs.items():
                # Sort groups by year for better organization
                groups.sort(key=lambda x: x['student_group']['name'])
                
                # Use main program name as sheet name, but limit length for Excel
                safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
                self.create_combined_program_sheet(wb, safe_sheet_name, groups)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SST_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"SST Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting SST timetables: {str(e)}"

    def export_tyd_timetables(self):
        """Export TYD (non-engineering) timetables"""
        data = self.load_saved_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Group TYD student groups by main program
            tyd_programs = defaultdict(list)
            
            for group_data in data:
                group_name = group_data['student_group']['name']
                if not self.is_sst_group(group_name):  # Not SST = TYD
                    main_program = self.extract_main_program_name(group_name)
                    tyd_programs[main_program].append(group_data)
            
            if not tyd_programs:
                return False, "No TYD student groups found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each unique program (one sheet per main program)
            for program_name, groups in tyd_programs.items():
                # Sort groups by year for better organization
                groups.sort(key=lambda x: x['student_group']['name'])
                
                # Use main program name as sheet name, but limit length for Excel
                safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
                self.create_combined_program_sheet(wb, safe_sheet_name, groups)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"TYD_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"TYD Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting TYD timetables: {str(e)}"

    def export_lecturer_timetables(self):
        """Export all lecturer timetables"""
        data = self.load_saved_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Collect all lecturer schedules
            lecturer_schedules = defaultdict(list)
            
            for group_data in data:
                student_group_name = group_data['student_group']['name']
                timetable_rows = group_data['timetable']
                
                for time_slot_idx, row_data in enumerate(timetable_rows):
                    for day_idx in range(len(self.days)):
                        if day_idx + 1 < len(row_data):  # Skip time column
                            cell_content = row_data[day_idx + 1]
                            lecturer_info = self.extract_lecturer_info(cell_content)
                            
                            if lecturer_info and lecturer_info['faculty']:
                                lecturer_name = lecturer_info['faculty']
                                
                                # Skip generic entries
                                if lecturer_name.lower() not in ['unknown', 'tbd', 'staff', '']:
                                    lecturer_schedules[lecturer_name].append({
                                        'time_slot': time_slot_idx,
                                        'day': day_idx,
                                        'course_code': lecturer_info['course_code'],
                                        'room': lecturer_info['room'],
                                        'student_group': student_group_name
                                    })
            
            if not lecturer_schedules:
                return False, "No lecturer data found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheet for each lecturer
            for lecturer_name, schedule in lecturer_schedules.items():
                self.create_lecturer_timetable_sheet(wb, lecturer_name, schedule)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Lecturer_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"Lecturer Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting lecturer timetables: {str(e)}"

# Main functions for easy import
def export_sst_timetables():
    """Export SST timetables"""
    exporter = TimetableExporter()
    return exporter.export_sst_timetables()

def export_tyd_timetables():
    """Export TYD timetables"""
    exporter = TimetableExporter()
    return exporter.export_tyd_timetables()

def export_lecturer_timetables():
    """Export lecturer timetables"""
    exporter = TimetableExporter()
    return exporter.export_lecturer_timetables()

if __name__ == "__main__":
    # Test the export functions
    print("Testing SST export...")
    success, message = export_sst_timetables()
    print(f"SST: {message}")
    
    print("\nTesting TYD export...")
    success, message = export_tyd_timetables()
    print(f"TYD: {message}")
    
    print("\nTesting Lecturer export...")
    success, message = export_lecturer_timetables()
    print(f"Lecturer: {message}")
